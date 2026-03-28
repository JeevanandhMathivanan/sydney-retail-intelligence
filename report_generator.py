import pandas as pd
import mysql.connector
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os
from config import DB_CONFIG


def get_connection():
    return mysql.connector.connect(**DB_CONFIG)


def run_query(sql):
    conn = get_connection()
    df   = pd.read_sql(sql, conn)
    conn.close()
    return df


# ── Shared styles ─────────────────────────────────────
DARK_BLUE    = "1F4E79"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
PALE_BLUE    = "DEEAF1"
DARK_GREEN   = "375623"
LIGHT_GREEN  = "E2EFDA"
DARK_RED     = "C00000"
LIGHT_RED    = "FFCCCC"
ORANGE       = "C55A11"
LIGHT_ORANGE = "FCE4D6"
PURPLE       = "7030A0"
LIGHT_PURPLE = "E2D9F3"
WHITE        = "FFFFFF"
LIGHT_GREY   = "F2F2F2"
DARK_GREY    = "595959"
YELLOW       = "FFD700"
LIGHT_YELLOW = "FFFACD"

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, text, bg=DARK_BLUE, fg=WHITE, size=11, bold=True, align="center"):
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = BORDER


def style_data(cell, align="left", bg=WHITE, bold=False, color=None, size=10):
    cell.font      = Font(name="Calibri", size=size, bold=bold,
                          color=color if color else "000000")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDER


def style_currency(cell, value, bg=WHITE, bold=False):
    cell.value         = value
    cell.number_format = '$#,##0.00'
    cell.font          = Font(name="Calibri", size=10, bold=bold)
    cell.fill          = PatternFill("solid", fgColor=bg)
    cell.alignment     = Alignment(horizontal="right", vertical="center")
    cell.border        = BORDER


def style_pct(cell, value, bg=WHITE):
    cell.value         = value / 100
    cell.number_format = '0.0%'
    cell.font          = Font(name="Calibri", size=10)
    cell.fill          = PatternFill("solid", fgColor=bg)
    cell.alignment     = Alignment(horizontal="center", vertical="center")
    cell.border        = BORDER


def add_title(ws, text, subtitle=None, bg=DARK_BLUE):
    ws.row_dimensions[1].height = 45
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column if ws.max_column > 1 else 10)}1")
    t = ws["A1"]
    t.value     = text
    t.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=bg)
    t.alignment = Alignment(horizontal="center", vertical="center")

    if subtitle:
        ws.row_dimensions[2].height = 22
        ws.merge_cells(f"A2:{get_column_letter(ws.max_column if ws.max_column > 1 else 10)}2")
        s = ws["A2"]
        s.value     = subtitle
        s.font      = Font(name="Calibri", italic=True, size=10, color=DARK_GREY)
        s.fill      = PatternFill("solid", fgColor=PALE_BLUE)
        s.alignment = Alignment(horizontal="center", vertical="center")


# ════════════════════════════════════════════════════
# FETCH ALL DATA FROM MYSQL
# ════════════════════════════════════════════════════

def fetch_all_insights():
    print("Fetching insights from MySQL...")

    store_revenue = run_query("""
        SELECT
            s.store_name,
            s.region,
            s.target_monthly,
            s.target_monthly * 24              AS target_2yr,
            COUNT(t.transaction_id)            AS transactions,
            ROUND(SUM(t.total_revenue), 2)     AS total_revenue,
            ROUND(SUM(t.profit), 2)            AS total_profit,
            ROUND(AVG(t.total_revenue), 2)     AS avg_transaction,
            ROUND(AVG(t.margin_pct), 1)        AS avg_margin,
            ROUND(SUM(t.total_revenue) /
                (s.target_monthly * 24) * 100, 1) AS target_achieved_pct
        FROM transactions t
        JOIN stores s ON t.store_id = s.store_id
        GROUP BY s.store_id, s.store_name, s.region,
                 s.target_monthly
        ORDER BY total_revenue DESC
    """)

    top_products = run_query("""
        SELECT
            p.name                             AS product_name,
            p.category,
            COUNT(t.transaction_id)            AS times_sold,
            SUM(t.quantity)                    AS units_sold,
            ROUND(SUM(t.total_revenue), 2)     AS total_revenue,
            ROUND(SUM(t.profit), 2)            AS total_profit,
            ROUND(AVG(t.margin_pct), 1)        AS avg_margin_pct
        FROM transactions t
        JOIN products p ON t.product_id = p.product_id
        GROUP BY p.product_id, p.name, p.category
        ORDER BY total_revenue DESC
        LIMIT 15
    """)

    monthly_trend = run_query("""
        SELECT
            month,
            year,
            COUNT(transaction_id)              AS transactions,
            ROUND(SUM(total_revenue), 2)       AS monthly_revenue,
            ROUND(SUM(profit), 2)              AS monthly_profit,
            ROUND(AVG(total_revenue), 2)       AS avg_transaction
        FROM transactions
        GROUP BY month, year
        ORDER BY year, month
    """)

    customer_segments = run_query("""
        SELECT
            c.segment,
            COUNT(DISTINCT t.customer_id)      AS unique_customers,
            COUNT(t.transaction_id)            AS transactions,
            ROUND(SUM(t.total_revenue), 2)     AS total_revenue,
            ROUND(AVG(t.total_revenue), 2)     AS avg_spend_per_visit,
            ROUND(SUM(t.total_revenue) /
                COUNT(DISTINCT t.customer_id), 2) AS revenue_per_customer,
            ROUND(AVG(t.margin_pct), 1)        AS avg_margin
        FROM transactions t
        JOIN customers c ON t.customer_id = c.customer_id
        GROUP BY c.segment
        ORDER BY total_revenue DESC
    """)

    category_perf = run_query("""
        SELECT
            p.category,
            COUNT(t.transaction_id)            AS transactions,
            SUM(t.quantity)                    AS units_sold,
            ROUND(SUM(t.total_revenue), 2)     AS total_revenue,
            ROUND(SUM(t.profit), 2)            AS total_profit,
            ROUND(AVG(t.margin_pct), 1)        AS avg_margin,
            ROUND(SUM(t.total_revenue) * 100.0 /
                (SELECT SUM(total_revenue) FROM transactions), 1
            )                                  AS revenue_share_pct
        FROM transactions t
        JOIN products p ON t.product_id = p.product_id
        GROUP BY p.category
        ORDER BY total_revenue DESC
    """)

    day_of_week = run_query("""
        SELECT
            day_of_week,
            COUNT(transaction_id)              AS transactions,
            ROUND(SUM(total_revenue), 2)       AS total_revenue,
            ROUND(AVG(total_revenue), 2)       AS avg_transaction,
            ROUND(SUM(profit), 2)              AS total_profit
        FROM transactions
        GROUP BY day_of_week
        ORDER BY FIELD(day_of_week,
            'Monday','Tuesday','Wednesday',
            'Thursday','Friday','Saturday','Sunday')
    """)

    payment_methods = run_query("""
        SELECT
            payment_method,
            COUNT(transaction_id)              AS transactions,
            ROUND(SUM(total_revenue), 2)       AS total_revenue,
            ROUND(AVG(total_revenue), 2)       AS avg_transaction,
            ROUND(COUNT(transaction_id) * 100.0 /
                (SELECT COUNT(*) FROM transactions), 1
            )                                  AS usage_pct
        FROM transactions
        GROUP BY payment_method
        ORDER BY transactions DESC
    """)

    quarterly = run_query("""
        SELECT
            year,
            quarter,
            COUNT(transaction_id)              AS transactions,
            ROUND(SUM(total_revenue), 2)       AS quarterly_revenue,
            ROUND(SUM(profit), 2)              AS quarterly_profit,
            ROUND(AVG(margin_pct), 1)          AS avg_margin
        FROM transactions
        GROUP BY year, quarter
        ORDER BY year, quarter
    """)

    top_suburbs = run_query("""
        SELECT
            c.suburb,
            COUNT(DISTINCT t.customer_id)      AS customers,
            COUNT(t.transaction_id)            AS transactions,
            ROUND(SUM(t.total_revenue), 2)     AS total_revenue,
            ROUND(AVG(t.total_revenue), 2)     AS avg_spend
        FROM transactions t
        JOIN customers c ON t.customer_id = c.customer_id
        GROUP BY c.suburb
        ORDER BY total_revenue DESC
        LIMIT 15
    """)

    loyalty = run_query("""
        SELECT
            CASE WHEN c.loyalty_member = 1
                 THEN 'Loyalty Member'
                 ELSE 'Non Member' END          AS membership,
            COUNT(DISTINCT t.customer_id)       AS customers,
            COUNT(t.transaction_id)             AS transactions,
            ROUND(SUM(t.total_revenue), 2)      AS total_revenue,
            ROUND(AVG(t.total_revenue), 2)      AS avg_spend_per_visit,
            ROUND(SUM(t.total_revenue) /
                COUNT(DISTINCT t.customer_id), 2) AS revenue_per_customer
        FROM transactions t
        JOIN customers c ON t.customer_id = c.customer_id
        GROUP BY c.loyalty_member
        ORDER BY total_revenue DESC
    """)

    print("All insights fetched successfully")
    return {
        "store_revenue":     store_revenue,
        "top_products":      top_products,
        "monthly_trend":     monthly_trend,
        "customer_segments": customer_segments,
        "category_perf":     category_perf,
        "day_of_week":       day_of_week,
        "payment_methods":   payment_methods,
        "quarterly":         quarterly,
        "top_suburbs":       top_suburbs,
        "loyalty":           loyalty,
    }


# ════════════════════════════════════════════════════
# BUILD EXCEL REPORT
# ════════════════════════════════════════════════════

def build_report(data):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ───────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = DARK_BLUE

    total_rev    = data["store_revenue"]["total_revenue"].sum()
    total_profit = data["store_revenue"]["total_profit"].sum()
    total_txn    = data["store_revenue"]["transactions"].sum()
    avg_margin   = data["category_perf"]["avg_margin"].mean()
    top_store    = data["store_revenue"].iloc[0]["store_name"]
    top_product  = data["top_products"].iloc[0]["product_name"]
    top_segment  = data["customer_segments"].iloc[0]["segment"]
    top_day      = data["day_of_week"].sort_values("total_revenue", ascending=False).iloc[0]["day_of_week"]
    top_suburb   = data["top_suburbs"].iloc[0]["suburb"]
    top_category = data["category_perf"].iloc[0]["category"]

    for col in range(1, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # Title
    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value     = f"AusMart Sydney — Executive Insights Report  |  Generated: {datetime.today().strftime('%d %B %Y')}"
    t.font      = Font(name="Calibri", bold=True, size=15, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 45

    ws1.merge_cells("A2:H2")
    s = ws1["A2"]
    s.value     = "Data period: January 2024 — December 2025  |  5 Sydney stores  |  7,602 transactions  |  1,000 customers  |  50 products"
    s.font      = Font(name="Calibri", italic=True, size=10, color=DARK_GREY)
    s.fill      = PatternFill("solid", fgColor=PALE_BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # KPI section label
    ws1.merge_cells("A4:H4")
    kpi_label = ws1["A4"]
    kpi_label.value     = "KEY PERFORMANCE INDICATORS"
    kpi_label.font      = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    kpi_label.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[4].height = 25

    kpis = [
        ("Total Revenue",        f"${total_rev:,.2f}",       DARK_BLUE,  LIGHT_BLUE,   "2 years of trading across all stores"),
        ("Total Profit",         f"${total_profit:,.2f}",    DARK_GREEN, LIGHT_GREEN,  f"{total_profit/total_rev*100:.1f}% profit margin"),
        ("Total Transactions",   f"{total_txn:,}",           PURPLE,     LIGHT_PURPLE, "Avg 10.5 per day"),
        ("Avg Transaction Value",f"${total_rev/total_txn:.2f}",ORANGE,   LIGHT_ORANGE, "Per customer visit"),
        ("Avg Profit Margin",    f"{avg_margin:.1f}%",       DARK_RED,   LIGHT_RED,    "Across all categories"),
        ("Top Performing Store", top_store,                   DARK_BLUE,  LIGHT_BLUE,   "Highest total revenue"),
        ("Best Selling Product", top_product,                 DARK_GREEN, LIGHT_GREEN,  "By total revenue"),
        ("Top Customer Segment", top_segment,                 PURPLE,     LIGHT_PURPLE, "Highest revenue contribution"),
        ("Peak Trading Day",     top_day,                     ORANGE,     LIGHT_ORANGE, "Highest average daily revenue"),
        ("Top Sydney Suburb",    top_suburb,                  DARK_RED,   LIGHT_RED,    "By customer spend"),
    ]

    row = 5
    for i, (label, value, dark, light, note) in enumerate(kpis):
        col = (i % 2) * 4 + 1
        if i % 2 == 0 and i > 0:
            row += 4

        # Label cell
        lc = ws1.cell(row=row, column=col, value=label)
        lc.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=dark)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = BORDER
        ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        ws1.row_dimensions[row].height = 20

        # Value cell
        vc = ws1.cell(row=row+1, column=col, value=value)
        vc.font      = Font(name="Calibri", bold=True, size=18, color=dark)
        vc.fill      = PatternFill("solid", fgColor=light)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = BORDER
        ws1.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+2)
        ws1.row_dimensions[row+1].height = 35

        # Note cell
        nc = ws1.cell(row=row+2, column=col, value=note)
        nc.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
        nc.fill      = PatternFill("solid", fgColor=PALE_BLUE)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border    = BORDER
        ws1.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+2)
        ws1.row_dimensions[row+2].height = 18

    print("Sheet 1: Executive Summary done")

    # ── Sheet 2: Store Performance ───────────────────
    ws2 = wb.create_sheet("Store Performance")
    ws2.sheet_properties.tabColor = MID_BLUE

    cols2   = ["Store", "Region", "Transactions", "Total Revenue", "Total Profit",
               "Avg Transaction", "Avg Margin %", "2yr Target", "Target Achieved %"]
    widths2 = [22, 14, 14, 16, 16, 16, 14, 16, 18]

    ws2.merge_cells(f"A1:{get_column_letter(len(cols2))}1")
    ws2.merge_cells(f"A2:{get_column_letter(len(cols2))}2")
    t2 = ws2["A1"]
    t2.value     = "AusMart Sydney — Store Performance Analysis"
    t2.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    t2.fill      = PatternFill("solid", fgColor=MID_BLUE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40
    s2 = ws2["A2"]
    s2.value     = "January 2024 — December 2025  |  Revenue, profit and target achievement by store"
    s2.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    s2.fill      = PatternFill("solid", fgColor=PALE_BLUE)
    s2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 18

    for col, (h, w) in enumerate(zip(cols2, widths2), 1):
        style_header(ws2.cell(row=3, column=col), h, bg=MID_BLUE)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[3].height = 30

    df2 = data["store_revenue"]
    for ri, (_, row_data) in enumerate(df2.iterrows(), 4):
        bg = PALE_BLUE if ri % 2 == 0 else WHITE
        achieved = row_data["target_achieved_pct"]
        achieved_bg = LIGHT_GREEN if achieved >= 100 else LIGHT_YELLOW if achieved >= 80 else LIGHT_RED

        style_data(ws2.cell(row=ri, column=1, value=row_data["store_name"]),  bold=True, bg=bg)
        style_data(ws2.cell(row=ri, column=2, value=row_data["region"]),       bg=bg)
        style_data(ws2.cell(row=ri, column=3, value=int(row_data["transactions"])), align="center", bg=bg)
        style_currency(ws2.cell(row=ri, column=4), row_data["total_revenue"],  bg=bg, bold=True)
        style_currency(ws2.cell(row=ri, column=5), row_data["total_profit"],   bg=bg)
        style_currency(ws2.cell(row=ri, column=6), row_data["avg_transaction"],bg=bg)
        style_pct(ws2.cell(row=ri, column=7),      row_data["avg_margin"],     bg=bg)
        style_currency(ws2.cell(row=ri, column=8), row_data["target_2yr"],     bg=bg)
        style_pct(ws2.cell(row=ri, column=9),      row_data["target_achieved_pct"], bg=achieved_bg)
        ws2.row_dimensions[ri].height = 22

    # Bar chart — revenue by store
    chart2 = BarChart()
    chart2.type    = "bar"
    chart2.title   = "Total Revenue by Store"
    chart2.y_axis.title = "Store"
    chart2.x_axis.title = "Revenue ($)"
    chart2.height  = 12
    chart2.width   = 20

    data_ref  = Reference(ws2, min_col=4, max_col=4, min_row=3, max_row=3+len(df2))
    cats_ref  = Reference(ws2, min_col=1, max_col=1, min_row=4, max_row=3+len(df2))
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    ws2.add_chart(chart2, f"A{5+len(df2)+2}")
    print("Sheet 2: Store Performance done")

    # ── Sheet 3: Product Analysis ─────────────────────
    ws3 = wb.create_sheet("Product Analysis")
    ws3.sheet_properties.tabColor = DARK_GREEN

    cols3   = ["Product", "Category", "Times Sold", "Units Sold",
               "Total Revenue", "Total Profit", "Avg Margin %"]
    widths3 = [30, 14, 12, 12, 16, 16, 14]

    ws3.merge_cells(f"A1:{get_column_letter(len(cols3))}1")
    ws3.merge_cells(f"A2:{get_column_letter(len(cols3))}2")
    t3 = ws3["A1"]
    t3.value     = "AusMart Sydney — Top 15 Products by Revenue"
    t3.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    t3.fill      = PatternFill("solid", fgColor=DARK_GREEN)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40
    s3 = ws3["A2"]
    s3.value     = "Ranked by total revenue — includes margin analysis and units sold"
    s3.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    s3.fill      = PatternFill("solid", fgColor=LIGHT_GREEN)
    s3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 18

    for col, (h, w) in enumerate(zip(cols3, widths3), 1):
        style_header(ws3.cell(row=3, column=col), h, bg=DARK_GREEN)
        ws3.column_dimensions[get_column_letter(col)].width = w

    df3 = data["top_products"]
    for ri, (_, row_data) in enumerate(df3.iterrows(), 4):
        bg       = LIGHT_GREEN if ri % 2 == 0 else WHITE
        rank_bg  = LIGHT_YELLOW if ri <= 6 else bg
        style_data(ws3.cell(row=ri, column=1, value=row_data["product_name"]), bold=ri<=6, bg=rank_bg)
        style_data(ws3.cell(row=ri, column=2, value=row_data["category"]),     bg=bg)
        style_data(ws3.cell(row=ri, column=3, value=int(row_data["times_sold"])),  align="center", bg=bg)
        style_data(ws3.cell(row=ri, column=4, value=int(row_data["units_sold"])),  align="center", bg=bg)
        style_currency(ws3.cell(row=ri, column=5), row_data["total_revenue"],  bg=rank_bg, bold=ri<=6)
        style_currency(ws3.cell(row=ri, column=6), row_data["total_profit"],   bg=bg)
        style_pct(ws3.cell(row=ri, column=7),      row_data["avg_margin_pct"],bg=bg)
        ws3.row_dimensions[ri].height = 22

    print("Sheet 3: Product Analysis done")

    # ── Sheet 4: Monthly Trend ────────────────────────
    ws4 = wb.create_sheet("Monthly Trend")
    ws4.sheet_properties.tabColor = ORANGE

    cols4   = ["Month", "Year", "Transactions", "Monthly Revenue", "Monthly Profit", "Avg Transaction"]
    widths4 = [12, 8, 14, 18, 18, 18]

    ws4.merge_cells(f"A1:{get_column_letter(len(cols4))}1")
    ws4.merge_cells(f"A2:{get_column_letter(len(cols4))}2")
    t4 = ws4["A1"]
    t4.value     = "AusMart Sydney — Monthly Revenue Trend"
    t4.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    t4.fill      = PatternFill("solid", fgColor=ORANGE)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 40
    s4 = ws4["A2"]
    s4.value     = "Month by month revenue, profit and transaction volume — January 2024 to December 2025"
    s4.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    s4.fill      = PatternFill("solid", fgColor=LIGHT_ORANGE)
    s4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[2].height = 18

    for col, (h, w) in enumerate(zip(cols4, widths4), 1):
        style_header(ws4.cell(row=3, column=col), h, bg=ORANGE)
        ws4.column_dimensions[get_column_letter(col)].width = w

    df4 = data["monthly_trend"]
    for ri, (_, row_data) in enumerate(df4.iterrows(), 4):
        bg = LIGHT_ORANGE if ri % 2 == 0 else WHITE
        style_data(ws4.cell(row=ri, column=1, value=row_data["month"]),            align="center", bg=bg)
        style_data(ws4.cell(row=ri, column=2, value=int(row_data["year"])),        align="center", bg=bg)
        style_data(ws4.cell(row=ri, column=3, value=int(row_data["transactions"])),align="center", bg=bg)
        style_currency(ws4.cell(row=ri, column=4), row_data["monthly_revenue"],    bg=bg, bold=True)
        style_currency(ws4.cell(row=ri, column=5), row_data["monthly_profit"],     bg=bg)
        style_currency(ws4.cell(row=ri, column=6), row_data["avg_transaction"],    bg=bg)
        ws4.row_dimensions[ri].height = 22

    # Line chart — monthly revenue
    chart4 = LineChart()
    chart4.title   = "Monthly Revenue Trend"
    chart4.y_axis.title = "Revenue ($)"
    chart4.x_axis.title = "Month"
    chart4.height  = 14
    chart4.width   = 24
    chart4.style   = 10

    rev_ref  = Reference(ws4, min_col=4, max_col=4, min_row=3, max_row=3+len(df4))
    prof_ref = Reference(ws4, min_col=5, max_col=5, min_row=3, max_row=3+len(df4))
    cats4    = Reference(ws4, min_col=1, max_col=1, min_row=4, max_row=3+len(df4))
    chart4.add_data(rev_ref,  titles_from_data=True)
    chart4.add_data(prof_ref, titles_from_data=True)
    chart4.set_categories(cats4)
    ws4.add_chart(chart4, f"A{5+len(df4)+2}")
    print("Sheet 4: Monthly Trend done")

    # ── Sheet 5: Customer Insights ────────────────────
    ws5 = wb.create_sheet("Customer Insights")
    ws5.sheet_properties.tabColor = PURPLE

    ws5.merge_cells("A1:H1")
    ws5.merge_cells("A2:H2")
    t5 = ws5["A1"]
    t5.value     = "AusMart Sydney — Customer Insights"
    t5.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    t5.fill      = PatternFill("solid", fgColor=PURPLE)
    t5.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 40
    s5 = ws5["A2"]
    s5.value     = "Segment analysis, loyalty performance and top Sydney suburbs"
    s5.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    s5.fill      = PatternFill("solid", fgColor=LIGHT_PURPLE)
    s5.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[2].height = 18

    # Section A — Segments
    ws5.cell(row=4, column=1).value = "CUSTOMER SEGMENTS"
    ws5.cell(row=4, column=1).font  = Font(name="Calibri", bold=True, size=11, color=PURPLE)
    ws5.row_dimensions[4].height = 22

    seg_headers = ["Segment", "Customers", "Transactions", "Total Revenue",
                   "Avg Spend/Visit", "Revenue/Customer", "Avg Margin %"]
    seg_widths  = [14, 12, 14, 18, 16, 18, 14]

    for col, (h, w) in enumerate(zip(seg_headers, seg_widths), 1):
        style_header(ws5.cell(row=5, column=col), h, bg=PURPLE)
        ws5.column_dimensions[get_column_letter(col)].width = w

    df5a = data["customer_segments"]
    for ri, (_, row_data) in enumerate(df5a.iterrows(), 6):
        bg = LIGHT_PURPLE if ri % 2 == 0 else WHITE
        style_data(ws5.cell(row=ri, column=1, value=row_data["segment"]),              bold=True, bg=bg)
        style_data(ws5.cell(row=ri, column=2, value=int(row_data["unique_customers"])),align="center", bg=bg)
        style_data(ws5.cell(row=ri, column=3, value=int(row_data["transactions"])),    align="center", bg=bg)
        style_currency(ws5.cell(row=ri, column=4), row_data["total_revenue"],          bg=bg, bold=True)
        style_currency(ws5.cell(row=ri, column=5), row_data["avg_spend_per_visit"],    bg=bg)
        style_currency(ws5.cell(row=ri, column=6), row_data["revenue_per_customer"],   bg=bg)
        style_pct(ws5.cell(row=ri, column=7),      row_data["avg_margin"],             bg=bg)
        ws5.row_dimensions[ri].height = 22

    # Section B — Loyalty
    loyalty_row = 6 + len(df5a) + 2
    ws5.cell(row=loyalty_row, column=1).value = "LOYALTY MEMBER ANALYSIS"
    ws5.cell(row=loyalty_row, column=1).font  = Font(name="Calibri", bold=True, size=11, color=PURPLE)
    ws5.row_dimensions[loyalty_row].height = 22

    loy_headers = ["Membership Status", "Customers", "Transactions",
                   "Total Revenue", "Avg Spend/Visit", "Revenue/Customer"]
    for col, h in enumerate(loy_headers, 1):
        style_header(ws5.cell(row=loyalty_row+1, column=col), h, bg=PURPLE)

    df5b = data["loyalty"]
    for ri, (_, row_data) in enumerate(df5b.iterrows(), loyalty_row+2):
        bg = LIGHT_PURPLE if ri % 2 == 0 else WHITE
        style_data(ws5.cell(row=ri, column=1, value=row_data["membership"]),           bold=True, bg=bg)
        style_data(ws5.cell(row=ri, column=2, value=int(row_data["customers"])),       align="center", bg=bg)
        style_data(ws5.cell(row=ri, column=3, value=int(row_data["transactions"])),    align="center", bg=bg)
        style_currency(ws5.cell(row=ri, column=4), row_data["total_revenue"],          bg=bg, bold=True)
        style_currency(ws5.cell(row=ri, column=5), row_data["avg_spend_per_visit"],    bg=bg)
        style_currency(ws5.cell(row=ri, column=6), row_data["revenue_per_customer"],   bg=bg)
        ws5.row_dimensions[ri].height = 22

    # Section C — Top suburbs
    suburb_row = loyalty_row + 2 + len(df5b) + 2
    ws5.cell(row=suburb_row, column=1).value = "TOP 15 SYDNEY SUBURBS BY REVENUE"
    ws5.cell(row=suburb_row, column=1).font  = Font(name="Calibri", bold=True, size=11, color=PURPLE)
    ws5.row_dimensions[suburb_row].height = 22

    sub_headers = ["Suburb", "Customers", "Transactions", "Total Revenue", "Avg Spend"]
    for col, h in enumerate(sub_headers, 1):
        style_header(ws5.cell(row=suburb_row+1, column=col), h, bg=PURPLE)

    df5c = data["top_suburbs"]
    for ri, (_, row_data) in enumerate(df5c.iterrows(), suburb_row+2):
        bg = LIGHT_PURPLE if ri % 2 == 0 else WHITE
        style_data(ws5.cell(row=ri, column=1, value=row_data["suburb"]),               bg=bg)
        style_data(ws5.cell(row=ri, column=2, value=int(row_data["customers"])),       align="center", bg=bg)
        style_data(ws5.cell(row=ri, column=3, value=int(row_data["transactions"])),    align="center", bg=bg)
        style_currency(ws5.cell(row=ri, column=4), row_data["total_revenue"],          bg=bg, bold=True)
        style_currency(ws5.cell(row=ri, column=5), row_data["avg_spend"],              bg=bg)
        ws5.row_dimensions[ri].height = 22

    print("Sheet 5: Customer Insights done")

    # ── Sheet 6: Category and Operations ─────────────
    ws6 = wb.create_sheet("Category & Operations")
    ws6.sheet_properties.tabColor = DARK_RED

    ws6.merge_cells("A1:H1")
    ws6.merge_cells("A2:H2")
    t6 = ws6["A1"]
    t6.value     = "AusMart Sydney — Category Performance & Operational Insights"
    t6.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    t6.fill      = PatternFill("solid", fgColor=DARK_RED)
    t6.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 40
    s6 = ws6["A2"]
    s6.value     = "Category revenue share, day-of-week patterns and payment method breakdown"
    s6.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    s6.fill      = PatternFill("solid", fgColor=LIGHT_RED)
    s6.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[2].height = 18

    # Category performance
    ws6.cell(row=4, column=1).value = "CATEGORY PERFORMANCE"
    ws6.cell(row=4, column=1).font  = Font(name="Calibri", bold=True, size=11, color=DARK_RED)

    cat_headers = ["Category", "Transactions", "Units Sold", "Total Revenue",
                   "Total Profit", "Avg Margin %", "Revenue Share %"]
    cat_widths  = [16, 14, 12, 18, 16, 14, 16]

    for col, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
        style_header(ws6.cell(row=5, column=col), h, bg=DARK_RED)
        ws6.column_dimensions[get_column_letter(col)].width = w

    df6a = data["category_perf"]
    for ri, (_, row_data) in enumerate(df6a.iterrows(), 6):
        bg = LIGHT_RED if ri % 2 == 0 else WHITE
        style_data(ws6.cell(row=ri, column=1, value=row_data["category"]),              bold=True, bg=bg)
        style_data(ws6.cell(row=ri, column=2, value=int(row_data["transactions"])),     align="center", bg=bg)
        style_data(ws6.cell(row=ri, column=3, value=int(row_data["units_sold"])),       align="center", bg=bg)
        style_currency(ws6.cell(row=ri, column=4), row_data["total_revenue"],           bg=bg, bold=True)
        style_currency(ws6.cell(row=ri, column=5), row_data["total_profit"],            bg=bg)
        style_pct(ws6.cell(row=ri, column=6),      row_data["avg_margin"],              bg=bg)
        style_pct(ws6.cell(row=ri, column=7),      row_data["revenue_share_pct"],       bg=bg)
        ws6.row_dimensions[ri].height = 22

    # Day of week
    dow_row = 6 + len(df6a) + 2
    ws6.cell(row=dow_row, column=1).value = "SALES BY DAY OF WEEK"
    ws6.cell(row=dow_row, column=1).font  = Font(name="Calibri", bold=True, size=11, color=DARK_RED)

    dow_headers = ["Day", "Transactions", "Total Revenue", "Avg Transaction", "Total Profit"]
    for col, h in enumerate(dow_headers, 1):
        style_header(ws6.cell(row=dow_row+1, column=col), h, bg=DARK_RED)

    df6b = data["day_of_week"]
    max_rev = df6b["total_revenue"].max()
    for ri, (_, row_data) in enumerate(df6b.iterrows(), dow_row+2):
        is_peak = row_data["total_revenue"] == max_rev
        bg      = LIGHT_YELLOW if is_peak else (LIGHT_RED if ri % 2 == 0 else WHITE)
        style_data(ws6.cell(row=ri, column=1, value=row_data["day_of_week"]),           bold=is_peak, bg=bg)
        style_data(ws6.cell(row=ri, column=2, value=int(row_data["transactions"])),     align="center", bg=bg)
        style_currency(ws6.cell(row=ri, column=3), row_data["total_revenue"],           bg=bg, bold=is_peak)
        style_currency(ws6.cell(row=ri, column=4), row_data["avg_transaction"],         bg=bg)
        style_currency(ws6.cell(row=ri, column=5), row_data["total_profit"],            bg=bg)
        ws6.row_dimensions[ri].height = 22

    # Payment methods
    pay_row = dow_row + 2 + len(df6b) + 2
    ws6.cell(row=pay_row, column=1).value = "PAYMENT METHOD BREAKDOWN"
    ws6.cell(row=pay_row, column=1).font  = Font(name="Calibri", bold=True, size=11, color=DARK_RED)

    pay_headers = ["Payment Method", "Transactions", "Total Revenue", "Avg Transaction", "Usage %"]
    for col, h in enumerate(pay_headers, 1):
        style_header(ws6.cell(row=pay_row+1, column=col), h, bg=DARK_RED)

    df6c = data["payment_methods"]
    for ri, (_, row_data) in enumerate(df6c.iterrows(), pay_row+2):
        bg = LIGHT_RED if ri % 2 == 0 else WHITE
        style_data(ws6.cell(row=ri, column=1, value=row_data["payment_method"]),        bg=bg)
        style_data(ws6.cell(row=ri, column=2, value=int(row_data["transactions"])),     align="center", bg=bg)
        style_currency(ws6.cell(row=ri, column=3), row_data["total_revenue"],           bg=bg, bold=True)
        style_currency(ws6.cell(row=ri, column=4), row_data["avg_transaction"],         bg=bg)
        style_pct(ws6.cell(row=ri, column=5),      row_data["usage_pct"],               bg=bg)
        ws6.row_dimensions[ri].height = 22

    print("Sheet 6: Category & Operations done")

    # ── Save ─────────────────────────────────────────
    os.makedirs("output", exist_ok=True)
    filepath = f"output/AusMart_Insights_Report_{datetime.today().strftime('%Y%m%d')}.xlsx"
    wb.save(filepath)
    print(f"\nReport saved: {filepath}")
    return filepath


def run():
    print("="*55)
    print("AusMart Sydney — Insights Report Generator")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*55)

    data     = fetch_all_insights()
    filepath = build_report(data)

    print("\n" + "="*55)
    print("Report complete! 6 sheets generated:")
    print("  1. Executive Summary   — KPI cards")
    print("  2. Store Performance   — revenue vs target")
    print("  3. Product Analysis    — top 15 products")
    print("  4. Monthly Trend       — line chart")
    print("  5. Customer Insights   — segments + suburbs")
    print("  6. Category & Ops      — category + day + payment")
    print(f"\nFile: {filepath}")
    print("="*55)


if __name__ == "__main__":
    run()