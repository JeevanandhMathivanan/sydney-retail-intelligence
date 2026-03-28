import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
from datetime import datetime


def create_excel_intake():
    print("Creating AusMart Excel intake file...")

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────
    DARK_BLUE  = "1F4E79"
    MID_BLUE   = "2E75B6"
    LIGHT_BLUE = "BDD7EE"
    PALE_BLUE  = "DEEAF1"
    GREEN      = "375623"
    LIGHT_GREEN = "E2EFDA"
    ORANGE     = "C55A11"
    LIGHT_ORANGE = "FCE4D6"
    WHITE      = "FFFFFF"
    LIGHT_GREY = "F2F2F2"
    DARK_GREY  = "404040"

    def header_style(cell, text, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
        cell.value     = text
        cell.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    def data_style(cell, align="left", bg=WHITE):
        cell.font      = Font(name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = Border(
            left=Side(style="thin"),  right=Side(style="thin"),
            top=Side(style="thin"),   bottom=Side(style="thin")
        )

    # ══════════════════════════════════════════════════
    # Sheet 1 — Sales Transactions Intake
    # ══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Sales Transactions"
    ws1.sheet_properties.tabColor = "1F4E79"
    ws1.row_dimensions[1].height  = 40
    ws1.row_dimensions[2].height  = 20

    # Title row
    ws1.merge_cells("A1:N1")
    title = ws1["A1"]
    title.value     = "AusMart Sydney — Daily Sales Transaction Intake Form"
    title.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Instruction row
    ws1.merge_cells("A2:N2")
    inst = ws1["A2"]
    inst.value     = "Instructions: Enter one transaction per row. All highlighted fields are mandatory. Date format: DD/MM/YYYY"
    inst.font      = Font(name="Calibri", italic=True, size=9, color=DARK_GREY)
    inst.fill      = PatternFill("solid", fgColor=PALE_BLUE)
    inst.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers
    headers = [
        "Transaction ID", "Sale Date", "Store ID", "Store Name",
        "Customer ID", "Product ID", "Product Name", "Category",
        "Quantity", "Unit Price ($)", "Total Revenue ($)",
        "Payment Method", "Staff ID", "Notes"
    ]
    col_widths = [15, 14, 10, 22, 13, 12, 25, 14, 10, 14, 17, 16, 10, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col)
        header_style(cell, header)
        ws1.column_dimensions[get_column_letter(col)].width = width

    # Sample data rows
    sample_data = [
        [7603, "28/03/2026", 1, "AusMart CBD",        101, 11, "Samsung 55 TV",       "Electronics", 1, 999.00, 999.00,  "Card",           4,  ""],
        [7604, "28/03/2026", 2, "AusMart Parramatta", 205, 1,  "Full Cream Milk 2L",  "Grocery",     3, 3.80,   11.40,   "Cash",           5,  ""],
        [7605, "28/03/2026", 3, "AusMart Chatswood",  312, 23, "Winter Jacket",       "Clothing",    1, 149.00, 149.00,  "Digital Wallet", 9,  ""],
        [7606, "28/03/2026", 4, "AusMart Bondi",      445, 32, "Vitamin C 500mg 60pk","Health",      2, 19.99,  39.98,   "Card",           13, ""],
        [7607, "28/03/2026", 5, "AusMart Penrith",    550, 42, "Frying Pan 28cm",     "Home",        1, 49.99,  49.99,   "Card",           17, ""],
    ]

    for row_idx, row_data in enumerate(sample_data, 4):
        bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            data_style(cell, "center" if col_idx in [1,3,5,6,9] else "left", bg)

    # Data validation — Payment Method dropdown
    dv_payment = DataValidation(
        type="list",
        formula1='"Card,Cash,Digital Wallet,Loyalty Points"',
        allow_blank=True,
        showDropDown=False
    )
    dv_payment.sqref = "L4:L1000"
    ws1.add_data_validation(dv_payment)

    # Data validation — Store ID dropdown
    dv_store = DataValidation(
        type="list",
        formula1='"1,2,3,4,5"',
        allow_blank=True,
        showDropDown=False
    )
    dv_store.sqref = "C4:C1000"
    ws1.add_data_validation(dv_store)

    # Freeze panes so headers stay visible when scrolling
    ws1.freeze_panes = "A4"

    # ══════════════════════════════════════════════════
    # Sheet 2 — Store Reference
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Store Reference")
    ws2.sheet_properties.tabColor = "375623"

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value     = "AusMart Sydney — Store Reference Data"
    t2.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    t2.fill      = PatternFill("solid", fgColor=GREEN)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    store_headers = ["Store ID", "Store Name", "Suburb", "Region", "Manager", "Monthly Target ($)"]
    store_widths  = [10, 22, 16, 14, 18, 18]

    for col, (h, w) in enumerate(zip(store_headers, store_widths), 1):
        cell = ws2.cell(row=2, column=col)
        header_style(cell, h, bg=GREEN)
        ws2.column_dimensions[get_column_letter(col)].width = w

    store_data = [
        [1, "AusMart CBD",        "Sydney CBD", "CBD",         "Sarah Chen",    120000],
        [2, "AusMart Parramatta", "Parramatta", "Western",     "James Wilson",  95000],
        [3, "AusMart Chatswood",  "Chatswood",  "North Shore", "Priya Patel",   88000],
        [4, "AusMart Bondi",      "Bondi",      "Eastern",     "Michael Brown", 82000],
        [5, "AusMart Penrith",    "Penrith",    "Western",     "Lisa Nguyen",   75000],
    ]

    for row_idx, row in enumerate(store_data, 3):
        bg = LIGHT_GREEN if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            data_style(cell, "center" if col_idx in [1] else "left", bg)

    # ══════════════════════════════════════════════════
    # Sheet 3 — Product Reference
    # ══════════════════════════════════════════════════
    ws3 = wb.create_sheet("Product Reference")
    ws3.sheet_properties.tabColor = "C55A11"

    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value     = "AusMart Sydney — Product Reference Data"
    t3.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    t3.fill      = PatternFill("solid", fgColor=ORANGE)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35

    prod_headers = ["Product ID", "Product Name", "Category", "Cost ($)", "Price ($)", "Margin %", "Markup ($)"]
    prod_widths  = [12, 28, 14, 10, 10, 10, 12]

    for col, (h, w) in enumerate(zip(prod_headers, prod_widths), 1):
        cell = ws3.cell(row=2, column=col)
        header_style(cell, h, bg=ORANGE)
        ws3.column_dimensions[get_column_letter(col)].width = w

    products = [
        [1,  "Full Cream Milk 2L",        "Grocery",     2.20, 3.80,  42, 1.60],
        [2,  "Sourdough Bread 750g",      "Grocery",     2.50, 5.50,  55, 3.00],
        [11, "Samsung 55 TV",             "Electronics", 650,  999,   35, 349],
        [13, "Wireless Earbuds",          "Electronics", 45,   99,    55, 54],
        [21, "Cotton T-Shirt",            "Clothing",    8,    29.99, 73, 21.99],
        [23, "Winter Jacket",             "Clothing",    45,   149,   70, 104],
        [31, "Shampoo 400ml",             "Health",      3.50, 9.99,  65, 6.49],
        [32, "Vitamin C 500mg 60pk",      "Health",      6,    19.99, 70, 13.99],
        [41, "Coffee Mug",                "Home",        4,    14.99, 73, 10.99],
        [42, "Frying Pan 28cm",           "Home",        18,   49.99, 64, 31.99],
    ]

    for row_idx, row in enumerate(products, 3):
        bg = LIGHT_ORANGE if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            data_style(cell, "center" if col_idx in [1, 6] else "left", bg)

    # ══════════════════════════════════════════════════
    # Sheet 4 — Summary Dashboard (static snapshot)
    # ══════════════════════════════════════════════════
    ws4 = wb.create_sheet("Summary Dashboard")
    ws4.sheet_properties.tabColor = "7030A0"

    PURPLE = "7030A0"
    LIGHT_PURPLE = "E2D9F3"

    ws4.merge_cells("A1:H1")
    t4 = ws4["A1"]
    t4.value     = f"AusMart Sydney — Executive Summary  |  Generated: {datetime.today().strftime('%d %B %Y')}"
    t4.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    t4.fill      = PatternFill("solid", fgColor=PURPLE)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 35

    # KPI section
    kpis = [
        ("Total Revenue",       "$764,064",  "2 years of trading"),
        ("Total Profit",        "$416,940",  "54.6% profit margin"),
        ("Transactions",        "7,602",     "Across 5 stores"),
        ("Avg Transaction",     "$100.50",   "Per customer visit"),
        ("Top Store",           "CBD Store", "Highest revenue"),
        ("Best Day",            "Saturday",  "Peak trading day"),
    ]

    ws4.cell(row=3, column=1).value = "KEY PERFORMANCE INDICATORS"
    ws4.cell(row=3, column=1).font  = Font(name="Calibri", bold=True, size=11, color=PURPLE)

    for idx, (label, value, note) in enumerate(kpis, 4):
        col = ((idx - 4) % 3) * 3 + 1
        row = 4 + ((idx - 4) // 3) * 3

        # Label
        lc = ws4.cell(row=row, column=col, value=label)
        lc.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=MID_BLUE)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        # Value
        vc = ws4.cell(row=row+1, column=col, value=value)
        vc.font      = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
        vc.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
        vc.alignment = Alignment(horizontal="center", vertical="center")

        # Note
        nc = ws4.cell(row=row+2, column=col, value=note)
        nc.font      = Font(name="Calibri", italic=True, size=8, color=DARK_GREY)
        nc.fill      = PatternFill("solid", fgColor=PALE_BLUE)
        nc.alignment = Alignment(horizontal="center", vertical="center")

        for r in [row, row+1, row+2]:
            ws4.merge_cells(start_row=r, start_column=col,
                            end_row=r,   end_column=col+1)
            ws4.row_dimensions[r].height = 22

    for col in range(1, 9):
        ws4.column_dimensions[get_column_letter(col)].width = 16

    # Save
    os.makedirs("output", exist_ok=True)
    filepath = "output/AusMart_Sales_Intake.xlsx"
    wb.save(filepath)
    print(f"Excel file saved: {filepath}")
    print("Sheets created:")
    print("  1. Sales Transactions — data entry form with dropdowns")
    print("  2. Store Reference    — store lookup table")
    print("  3. Product Reference  — product catalog")
    print("  4. Summary Dashboard  — executive KPI snapshot")
    return filepath


if __name__ == "__main__":
    print("="*55)
    print("AusMart Sydney — Excel Intake Generator")
    print("="*55)
    create_excel_intake()
    print("="*55)
    print("Done! Open output/AusMart_Sales_Intake.xlsx")
    print("="*55)